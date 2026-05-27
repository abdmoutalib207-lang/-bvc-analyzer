# ============================================================
# BVC ANALYZER v5.0 — FUSION COMPLETE
# Base: v3.4 (architecture OO) + v3.5.3 (PDF, validation, backtest, exports)
# + v4.0 PRO (ponderation dynamique, EVA, liquidite, ACHAT FORT)
# ============================================================
# Cellule 1: INSTALLATION (a executer separement si besoin)
# !pip install pdfplumber requests openpyxl pandas numpy -q
# ============================================================

import subprocess, sys, os, re, json, time, base64, logging, warnings
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Any, Tuple, Union
from enum import Enum
from pathlib import Path
from io import BytesIO

import pandas as pd
import numpy as np
import requests
import pdfplumber

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo
    TZ_CASABLANCA = ZoneInfo("Africa/Casablanca")
except ImportError:
    TZ_CASABLANCA = None

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-8s | %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("BVC")

# ============================================================
# PARAMETRES CONFIGURABLES
# ============================================================

GITHUB_TOKEN = ""  # optionnel — laisser vide pour ne pas pusher
GITHUB_URL = "https://raw.githubusercontent.com/abdmoutalib207-lang/-bvc-analyzer/main/fondamentaux.json"
GITHUB_RAW_FOND = GITHUB_URL

OUTPUT_DIR = "/content"
FOND_PATH = "/content/fondamentaux.json"

# Frais reels BVC
BUY_FEE = 0.01       # 1% commission achat
SELL_FEE = 0.01      # 1% commission vente
TAX_RATE = 0.15      # 15% taxe plus-value
ATR_MULT = 2.5       # Stop = entree - ATR_MULT * ATR(14)
MIN_VOL_DH = 50_000  # Volume min en DH pour liquidite
SPLIT_EXCL = 5       # Jours a exclure post-split

HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Accept-Language": "fr-FR,fr;q=0.9"
}

# Sources PDF
PDF_SOURCES = {
    "SGTM": "https://attijaricib.com/sites/default/files/2025-12/IPO%20SGTM-Nov.%202025.pdf",
    "TGC":  "https://www.cdgcapital.ma/sites/default/files/publication/2021-12/Introduction%20en%20Bourse%20du%20groupe%20TGCC.pdf",
    "CASH": "https://attijaricib.com/sites/default/files/2025-11/IPO%20Cash%20Plus-%20Novembre%202025.pdf",
    "CSR":  "https://attijaricib.com/sites/default/files/2025-03/Research%20Report%20-%20Cosumar%20Mars%202025.pdf",
    "MSA":  "https://attijaricib.com/sites/default/files/2024-11/Research%20Report%20-%20Marsa%20Maroc%202024_0.pdf",
    "VCNE": "https://media.casablanca-bourse.com/sites/default/files/2025-06/ipo_vicenne_no_023_2025_0.pdf",
    "CFGB": "https://attijaricib.com/sites/default/files/2023-12/AGR-IPO%20CFG%202023_0.pdf",
    "MNG":  "https://attijaricib.com/sites/default/files/2024-04/AGR-CI%20Managem%202024.pdf",
    "CMGP": "https://attijaricib.com/sites/default/files/2024-11/IPO%20CMGP%20Group%202024.pdf",
}

CHAMPS_OBLIGATOIRES = [
    "secteur", "per", "croissance_ca", "dette_nette_ebitda",
    "roic", "wacc", "marge_nette", "cash_conversion", "div_yield",
    "momentum_fondamental", "catalyseur", "risque", "source", "date_maj"
]

# Dossiers
for folder in ["data/historique", "exports", OUTPUT_DIR]:
    Path(folder).mkdir(parents=True, exist_ok=True)

# ============================================================
# ISIN MAP — 80 TITRES + 5 IPO
# ============================================================

ISIN_MAP = {
    "ADH": "MA0000011512", "AFM": "MA0000012296", "AFI": "MA0000012114",
    "GAZ": "MA0000010951", "AGM": "MA0000010944", "AKD": "MA0000012585",
    "ADI": "MA0000011819", "ALU": "MA0000010936", "ARD": "MA0000012460",
    "ATL": "MA0000011710", "ATW": "MA0000012445", "HAL": "MA0000010969",
    "NEJ": "MA0000011009", "BAL": "MA0000011991", "BOA": "MA0000012437",
    "BCP": "MA0000011884", "BMC": "MA0000010811", "CAR": "MA0000011868",
    "CDM": "MA0000010381", "CIH": "MA0000011454", "CIM": "MA0000010506",
    "CMT": "MA0000011793", "COL": "MA0000011934", "CSR": "MA0000012247",
    "CTM": "MA0000010340", "DAR": "MA0000011421", "DHO": "MA0000011850",
    "DTT": "MA0000012536", "DSW": "MA0000011637", "ENK": "MA0000011942",
    "EQD": "MA0000010357", "FNB": "MA0000011587", "HPS": "MA0000011611",
    "IBM": "MA0000011132", "IMI": "MA0000012387", "INV": "MA0000011579",
    "JET": "MA0000012080", "LBV": "MA0000011801", "LHM": "MA0000012320",
    "LES": "MA0000012031", "M2M": "MA0000011678", "MOX": "MA0000010985",
    "MGL": "MA0000011215", "MNG": "MA0000011058", "MRL": "MA0000010035",
    "IAM": "MA0000011488", "MIC": "MA0000012163", "MUT": "MA0000012395",
    "OUL": "MA0000010415", "PPM": "MA0000011660", "REB": "MA0000010993",
    "RDS": "MA0000012239", "RIS": "MA0000011462", "S2M": "MA0000012106",
    "SLM": "MA0000012007", "SAF": "MA0000011744", "SMI": "MA0000010068",
    "STK": "MA0000011843", "SNP": "MA0000011728", "MSA": "MA0000012312",
    "SON": "MA0000010019", "SOT": "MA0000012502", "SRM": "MA0000011595",
    "SBS": "MA0000010365", "STR": "MA0000012056", "TQA": "MA0000012205",
    "TGC": "MA0000012528", "TIM": "MA0000011686", "TMA": "MA0000012262",
    "UNI": "MA0000012023", "WAF": "MA0000010928", "ZLD": "MA0000010571",
    # IPOs
    "CFGB": "MA0000012627", "CMGP": "MA0000012718", "VCNE": "MA0000012759",
    "CASH": "MA0000012767", "SGTM": "MA0000012783",
}

TICKERS_DEFAUT = [
    "SMI", "MNG", "RIS", "RDS", "CSR", "SOT",
    "ADI", "MSA", "ADH", "TGC", "CMT", "SRM",
    "CFGB", "CMGP", "VCNE", "CASH", "SGTM"
]


# ============================================================
# ENUMS ET STRUCTURES
# ============================================================

class Signal(Enum):
    ACHAT_FORT = "ACHAT FORT"
    ACHETER    = "ACHETER"
    ACCUMULER  = "ACCUMULER"
    SURVEILLER = "SURVEILLER"
    ATTENDRE   = "ATTENDRE"
    EVITER     = "EVITER"

class Setup(Enum):
    MOMENTUM_CONFIRME  = "MOMENTUM CONFIRME"
    BREAKOUT_POTENTIEL = "BREAKOUT POTENTIEL"
    PULLBACK_HAUSSIER  = "PULLBACK HAUSSIER"
    CONTRARIEN         = "CONTRARIEN"
    ACCUMULATION_LENTE = "ACCUMULATION LENTE"
    FAIBLESSE          = "FAIBLESSE"
    NEUTRE             = "NEUTRE"

class Severity(Enum):
    CRITIQUE = "CRITIQUE"
    ELEVEE   = "ELEVEE"
    MOYENNE  = "MOYENNE"
    FAIBLE   = "FAIBLE"

@dataclass
class RedFlag:
    severity: Severity
    category: str
    message: str
    is_blocking: bool = False

@dataclass
class ScoreResult:
    score: float
    max_score: float
    reasons: List[str]
    raw_details: Dict[str, Any] = field(default_factory=dict)
    @property
    def normalized(self):
        if self.max_score <= 0: return 0
        return round(min(10, self.score / self.max_score * 10), 2)

@dataclass
class Valuation:
    base: Optional[float]
    bull: Optional[float]
    bear: Optional[float]
    upside_base: Optional[float]
    upside_bull: Optional[float]
    downside: Optional[float]
    premium_pct: Optional[float]
    methodology: str
    confidence: str

@dataclass
class AnalysisResult:
    ticker: str
    price: float
    date: datetime
    setup: Setup
    signal: Signal
    action: str
    score_technical: ScoreResult
    score_fundamental: ScoreResult
    score_global: float
    red_flags: List[RedFlag]
    valuation: Valuation
    technical_data: Dict[str, Any]
    fundamental_data: Dict[str, Any]
    live_data: Dict[str, Any] = field(default_factory=dict)
    market_status: str = "UNKNOWN"
    institutional_note: str = ""
    quality_score: Dict[str, Any] = field(default_factory=dict)

# ============================================================
# SAFE MATH
# ============================================================

class SafeMath:
    @staticmethod
    def to_float(value):
        if value is None: return np.nan
        if isinstance(value, (list, dict, tuple)): return np.nan
        try:
            if pd.isna(value): return np.nan
        except: pass
        if isinstance(value, str):
            value = value.replace("\xa0"," ").replace(" ","").replace(",",".").replace("%","").strip()
            if value in ["","-","N/A","nan","None","—"]: return np.nan
        try: return float(value)
        except: return np.nan

    @staticmethod
    def last(series):
        s = series.dropna()
        return s.iloc[-1] if not s.empty else np.nan

    @staticmethod
    def prev(series, n=2):
        s = series.dropna()
        return s.iloc[-n] if len(s) >= n else np.nan

    @staticmethod
    def pct_change(a, b):
        if b == 0 or pd.isna(a) or pd.isna(b): return np.nan
        return round((a/b - 1)*100, 2)

# ============================================================
# CALENDRIER MARCHE
# ============================================================

class MarketCalendar:
    @staticmethod
    def now():
        if TZ_CASABLANCA:
            return datetime.now(TZ_CASABLANCA)
        return datetime.utcnow() + timedelta(hours=1)

    @staticmethod
    def status():
        now = MarketCalendar.now()
        if now.weekday() >= 5: return "WEEKEND"
        open_t  = now.replace(hour=9,  minute=30, second=0, microsecond=0)
        close_t = now.replace(hour=15, minute=30, second=0, microsecond=0)
        post_t  = now.replace(hour=15, minute=45, second=0, microsecond=0)
        if now < open_t:             return "PRE_MARKET"
        if open_t <= now <= close_t: return "OPEN"
        if close_t < now <= post_t:  return "POST_CLOSE"
        return "CLOSED"

    @staticmethod
    def message():
        s   = MarketCalendar.status()
        now = MarketCalendar.now()
        h   = now.strftime("%H:%M")
        return {
            "OPEN":       f"Marche OUVERT | Casablanca : {h}",
            "POST_CLOSE": f"Post-cloture | Casablanca : {h}",
            "PRE_MARKET": f"Avant ouverture | Casablanca : {h}",
            "WEEKEND":    f"Weekend | Casablanca : {h}",
            "CLOSED":     f"Marche ferme | Casablanca : {h}"
        }.get(s, f"Casablanca : {h}")


# ============================================================
# CONNECTEURS MARCHE (Medias24 + IDBourse fallback)
# ============================================================

class BVCLiveConnector:
    """Priorite: IDBourse -> Medias24 -> Cache"""

    IDB_BASE = "https://www.idbourse.com"
    MED_BASE = "https://medias24.com/content/api"

    TICKER_MAP = {
        "SMI":"SMI","MNG":"MNG","RIS":"RIS","RDS":"RDS","CSR":"CSR","SOT":"SOT",
        "ADI":"ALLI","MSA":"MSA","ADH":"ADH","TGC":"TGC","CMT":"MTO","SRM":"SRM",
        "CFGB":"CFG","CMGP":"CMGP","VCNE":"VCNE","CASH":"CASH","SGTM":"GTM",
    }

    _cache = None
    _cache_ts = 0
    _TTL = 300  # 5 minutes

    def __init__(self, isin_map: dict):
        self.isin_map = isin_map
        self.med_available = False
        self.idb_available = False

    def _idb_get(self, ep, timeout=10):
        try:
            hdrs = {**HEADERS_HTTP, "Referer": self.IDB_BASE}
            r = requests.get(f"{self.IDB_BASE}{ep}", headers=hdrs, timeout=timeout)
            if r.status_code == 200: return r.json()
        except: pass
        return None

    def _med_get(self, params: dict, timeout=10):
        try:
            r = requests.get(self.MED_BASE, params={**params, "format": "json"}, 
                           headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
            if r.status_code == 200 and len(r.text) > 50: return json.loads(r.text)
        except Exception as e: logger.debug(f"Medias24 erreur: {e}")
        return None

    def preload(self):
        # Test IDBourse
        m = self._idb_get("/api/proxy/masi-data")
        if m and m.get("value"):
            self.idb_available = True
            logger.info("IDBourse: disponible")
        # Test Medias24
        test = self._med_get({"method": "getStockInfo", "ISIN": "MA0000010068"}, timeout=8)
        if test and test.get("result"):
            self.med_available = True
            logger.info("Medias24: disponible")
        return self.idb_available or self.med_available

    def get_all_data(self, force=False):
        now = time.time()
        if not force and BVCLiveConnector._cache and (now - BVCLiveConnector._cache_ts) < self._TTL:
            return BVCLiveConnector._cache
        data = self._idb_get("/api/proxy/get_all_data")
        if data:
            BVCLiveConnector._cache = {d["name"]: d for d in data if "name" in d}
            BVCLiveConnector._cache_ts = now
            return BVCLiveConnector._cache
        return {}

    def get_stock_idb(self, ticker):
        sym = self.TICKER_MAP.get(ticker, ticker)
        return self._idb_get(f"/api/proxy/stock/{sym}")

    def get_cours(self, ticker):
        # 1. IDBourse direct
        s = self.get_stock_idb(ticker)
        if s and s.get("dernier_cours"):
            return float(s["dernier_cours"]), "idbourse"
        # 2. IDBourse cache
        sym = self.TICKER_MAP.get(ticker, ticker)
        for nom, d in self.get_all_data().items():
            if sym.upper() in nom.upper() or ticker.upper() in nom.upper():
                if d.get("dernier_cours"):
                    return float(d["dernier_cours"]), "idbourse_cache"
        # 3. Medias24
        if self.med_available:
            isin = self.isin_map.get(ticker.upper())
            if isin:
                data = self._med_get({"method": "getStockInfo", "ISIN": isin})
                if data and data.get("result"):
                    r = data["result"]
                    cours = SafeMath.to_float(r.get("cours"))
                    if cours: return cours, "medias24"
        return None, "indisponible"

    def get_historique(self, ticker: str, days: int = 730) -> pd.DataFrame:
        isin = self.isin_map.get(ticker.upper())
        if not isin:
            logger.warning(f"{ticker}: ISIN introuvable")
            return pd.DataFrame()

        end   = datetime.now().strftime("%Y-%m-%d")
        start = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

        data = self._med_get({"method": "getPriceHistory", "ISIN": isin, "from": start, "to": end})
        if not data or not data.get("result"):
            return pd.DataFrame()

        records = data["result"]
        df = pd.DataFrame({
            "date":   [r.get("date") for r in records],
            "close":  [SafeMath.to_float(r.get("value")) for r in records],
            "high":   [SafeMath.to_float(r.get("max")) for r in records],
            "low":    [SafeMath.to_float(r.get("min")) for r in records],
            "volume": [SafeMath.to_float(r.get("volume", 0)) for r in records],
        })
        df["date"] = pd.to_datetime(df["date"], dayfirst=True, errors="coerce")
        df["open"] = df["close"].shift(1).fillna(df["close"])
        df = df.dropna(subset=["date","close","high","low"])
        df = df.sort_values("date").reset_index(drop=True)
        df = self._adjust_splits(df, ticker)
        logger.info(f"Medias24 {ticker}: {len(df)} lignes | {df['close'].iloc[-1]} DH")
        return df

    def get_live(self, ticker: str) -> dict:
        isin = self.isin_map.get(ticker.upper())
        if not isin or not self.med_available: return {}
        data = self._med_get({"method": "getStockInfo", "ISIN": isin})
        if not data or not data.get("result"): return {}
        r = data["result"]
        return {
            "cours":      SafeMath.to_float(r.get("cours")),
            "ouverture":  SafeMath.to_float(r.get("ouverture")),
            "variation":  SafeMath.to_float(r.get("variation")),
            "volume_mad": SafeMath.to_float(r.get("volume")),
            "quantite":   SafeMath.to_float(r.get("volumeTitre")),
            "haut":       SafeMath.to_float(r.get("max")),
            "bas":        SafeMath.to_float(r.get("min")),
            "source":     "Medias24",
            "timestamp":  MarketCalendar.now().strftime("%H:%M:%S")
        }

    def _adjust_splits(self, df: pd.DataFrame, ticker: str) -> pd.DataFrame:
        df = df.copy().sort_values("date").reset_index(drop=True)
        rets = df["close"].pct_change()
        splits = rets[rets < -0.4].index.tolist()
        if splits:
            for idx in splits:
                ratio = df.loc[idx, "close"] / df.loc[idx-1, "close"]
                logger.info(f"Split detecte {ticker} a {df.loc[idx,'date'].date()} — ratio {ratio:.4f}")
                for col in ["close","open","high","low"]:
                    df.loc[:idx-1, col] = df.loc[:idx-1, col] * ratio
        return df

    def status(self):
        m = self._idb_get("/api/proxy/masi-data")
        if m and m.get("value"):
            v = m["value"]; var = m.get("variation", 0); cap = m.get("capitalisation", 0)/1e9
            logger.info(f"IDBourse OK — MASI: {v:,.2f}pts ({var:+.2f}%) | Cap: {cap:.1f}Mds DH")
            return True
        logger.warning("IDBourse indisponible")
        return False

# ============================================================
# HISTORIQUE LOCAL
# ============================================================

class HistoryStore:
    def __init__(self, base_path="data/historique"):
        self.base = Path(base_path)
        self.base.mkdir(parents=True, exist_ok=True)

    def path(self, ticker):
        return self.base / f"{ticker.upper()}.xlsx"

    def load(self, ticker):
        p = self.path(ticker)
        if not p.exists(): return pd.DataFrame()
        df = pd.read_excel(p)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        return df.dropna(subset=["date"])

    def save(self, ticker, df):
        df = df.copy()
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"])
        df = df.sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
        df.to_excel(self.path(ticker), index=False)
        return df

# ============================================================
# AUTO-RECUPERATION JSON
# ============================================================

def ensure_fondamentaux_json():
    if os.path.exists(FOND_PATH):
        logger.info("fondamentaux.json trouve localement")
        return FOND_PATH
    logger.warning("fondamentaux.json absent -> telechargement depuis GitHub...")
    try:
        r = requests.get(GITHUB_RAW_FOND, headers=HEADERS_HTTP, timeout=20)
        if r.status_code == 200 and len(r.text) > 100:
            with open(FOND_PATH, "w", encoding="utf-8") as f:
                f.write(r.text)
            logger.info("fondamentaux.json telecharge depuis GitHub")
            return FOND_PATH
    except Exception as e:
        logger.error(f"Erreur telechargement: {e}")
    raise FileNotFoundError("Impossible de recuperer fondamentaux.json")


# ============================================================
# MODULE EXTRACTION PDF (v3.5.3)
# ============================================================

class FundamentalPDFExtractor:
    CHAMP_MAP = [
        (["chiffre d'affaires","ca ","revenus","turnover"], "ca_mdh"),
        (["resultat net","rnpg","benefice net","net income"], "rn_mdh"),
        (["ebitda","ebe ","excedent brut"], "ebitda_mdh"),
        (["bpa","benefice par action","eps"], "bpa"),
        (["dpa","dividende par action","dps"], "dpa"),
        (["p/e","per ","price earning"], "per"),
        (["marge nette","net margin"], "marge_nette_pct"),
        (["marge ebitda","marge ebe"], "marge_ebitda_pct"),
        (["dette nette/ebitda","gearing","levier"], "dette_nette_ebitda"),
        (["rendement","dividend yield","d/y"], "div_yield_pct"),
        (["roic"], "roic_pct"),
    ]

    SEUILS = {
        "ca_mdh": (1, 500_000), "rn_mdh": (-10_000, 100_000),
        "ebitda_mdh": (0, 200_000), "bpa": (-500, 50_000),
        "dpa": (0, 10_000), "per": (0, 200),
        "marge_nette_pct": (-50, 80), "marge_ebitda_pct": (0, 80),
        "dette_nette_ebitda": (-5, 20), "div_yield_pct": (0, 30),
        "roic_pct": (-20, 100),
    }

    def __init__(self, delay=2.0):
        self.delay = delay
        self.log = []
        self.resultats = {}

    def _num(self, texte):
        if not texte: return None
        t = re.sub(r"[^\d,.\-]", "", str(texte).strip())
        if not t or t in ["-",".",","]: return None
        if "," in t and "." not in t:
            t = t.replace(",", ".")
        elif t.count(".") > 1:
            t = t.replace(".", "")
        try: return float(t)
        except: return None

    def _valider_champ(self, champ, valeur):
        if valeur is None: return False
        if champ not in self.SEUILS: return True
        lo, hi = self.SEUILS[champ]
        return lo <= valeur <= hi

    def _detecter_champ(self, label):
        label = label.lower().strip()
        for mots, champ in self.CHAMP_MAP:
            if any(m in label for m in mots): return champ
        return None

    def _get_pdf(self, url):
        verify = False if "casablanca-bourse" in url else True
        try:
            r = requests.get(url, headers=HEADERS_HTTP, timeout=40, verify=verify)
            if r.status_code == 200 and len(r.content) > 10_000: return r.content
        except Exception as e: logger.warning(f"Download PDF: {e}")
        return None

    def _pages_texte(self, pdf_bytes):
        pages = []
        try:
            with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
                for i, page in enumerate(pdf.pages, 1):
                    pages.append((i, page.extract_text() or ""))
        except: pass
        return pages

    def _parser_texte(self, pages_texte, ticker, url):
        histo = {}; traces = []; annees_ctx = []
        for page_num, texte in pages_texte:
            for ligne in texte.split("\n"):
                found = re.findall(r"(20\d{2}[Ee]?)", ligne)
                if len(found) >= 2:
                    annees_ctx = []
                    for a in found:
                        base = re.match(r"(20\d{2})", a).group(1)
                        suf = "E" if "E" in a or "e" in a else ""
                        annees_ctx.append(base + suf)
                    continue
                if not annees_ctx: continue
                champ = self._detecter_champ(ligne)
                if not champ: continue
                valeurs_raw = re.findall(r"[\-]?\d[\d\s,.]*", ligne)
                valeurs = [self._num(v) for v in valeurs_raw]
                valeurs = [v for v in valeurs if v is not None]
                for j, annee in enumerate(annees_ctx):
                    if j >= len(valeurs): break
                    val = valeurs[j]
                    if not self._valider_champ(champ, val): continue
                    histo.setdefault(annee, {})[champ] = val
                    traces.append({
                        "ticker": ticker, "source": url.split("/")[-1][:50],
                        "page": page_num, "methode": "texte",
                        "annee": annee, "champ": champ, "valeur": val,
                        "confiance": "moyenne", "ligne": ligne.strip()[:100],
                    })
        return histo, traces

    def _parser_meta(self, texte, url):
        meta = {}
        m = re.search(r"flottant[^\d]*(\d+[,]?\d*)\s*%", texte, re.IGNORECASE)
        if m:
            v = self._num(m.group(1))
            if v and 0 < v < 100: meta["free_float"] = v
        for rec in ["Souscrire","Acheter","Achat","Buy","Accumuler","Conserver","Neutre","Hold","Vendre","Sell"]:
            if re.search(r"\b" + rec + r"\b", texte, re.IGNORECASE):
                meta["recommandation"] = rec; break
        for src in ["Attijari Global Research","AGR","BKGR","BMCE Capital","CDG Capital","CFG Bank"]:
            if src.lower() in texte.lower():
                meta["analyste"] = src; break
        m = re.search(r"(?:cours cible|objectif|target)[^\d]*(\d[\d\s,.]+)\s*(?:DH|MAD)", texte, re.IGNORECASE)
        if m:
            v = self._num(m.group(1))
            if v and 10 < v < 100_000: meta["cours_cible"] = v
        meta["source_pdf"] = url.split("/")[-1][:80]
        return meta

    def analyser(self, ticker, url):
        logger.info(f"Analyse PDF {ticker}")
        pdf = self._get_pdf(url)
        if not pdf:
            r = {"ticker": ticker, "url": url, "erreur": "PDF non disponible",
                 "histo": {}, "meta": {}, "traces": []}
            self.resultats[ticker] = r; return r
        pages_texte = self._pages_texte(pdf)
        texte_complet = "\n".join(t for _, t in pages_texte)
        histo, traces = self._parser_texte(pages_texte, ticker, url)
        meta = self._parser_meta(texte_complet, url)
        self.log.extend(traces)
        logger.info(f"  {len(histo)} annees | {sum(len(v) for v in histo.values())} champs")
        r = {"ticker": ticker, "url": url, "erreur": None,
             "histo": histo, "meta": meta, "traces": traces}
        self.resultats[ticker] = r; return r

    def extraire_tous(self, sources=None):
        sources = sources or PDF_SOURCES
        logger.info(f"FundamentalPDFExtractor — {len(sources)} PDFs")
        for ticker, url in sources.items():
            self.analyser(ticker, url)
            time.sleep(self.delay)
        return self.resultats

    def enrichir_json(self, fond_path, output_path=None):
        with open(fond_path, "r", encoding="utf-8") as f:
            fond = json.load(f)
        for ticker, data in self.resultats.items():
            if ticker not in fond or data.get("erreur"): continue
            meta = data.get("meta", {}); histo = data.get("histo", {})
            if meta.get("free_float") is not None:
                fond[ticker]["free_float"] = meta["free_float"]
            if meta.get("cours_cible"):
                fond[ticker].setdefault("consensus_analystes", [])
                fond[ticker]["consensus_analystes"].append({
                    "cours_cible": meta.get("cours_cible"),
                    "recommandation": meta.get("recommandation"),
                    "analyste": meta.get("analyste"),
                    "source_pdf": meta.get("source_pdf"),
                    "date": datetime.today().strftime("%Y-%m-%d")
                })
            if histo:
                fond[ticker]["historique_previsionnel"] = histo
                fond[ticker].update(_tendances(histo))
            fond[ticker]["date_maj"] = datetime.today().strftime("%Y-%m-%d")
        output = output_path or os.path.join(OUTPUT_DIR, "fondamentaux_enrichi.json")
        with open(output, "w", encoding="utf-8") as f:
            json.dump(fond, f, ensure_ascii=False, indent=2)
        logger.info(f"fondamentaux_enrichi.json cree: {output}")
        return fond, output

def _tendances(histo):
    annees = sorted(a for a in histo if re.match(r"20\d{2}", a))
    def t(champ):
        vals = [histo[a].get(champ) for a in annees[-3:]]
        vals = [v for v in vals if v is not None]
        if len(vals) < 2: return "stable"
        base = vals[0] if vals[0] else 1
        d = (vals[-1] - vals[0]) / abs(base)
        if d > 0.05: return "hausse"
        if d < -0.05: return "baisse"
        return "stable"
    return {
        "tendance_roic": t("roic_pct"),
        "tendance_marge": t("marge_nette_pct"),
        "tendance_dette": t("dette_nette_ebitda"),
    }

# ============================================================
# MODULE VALIDATION JSON (v3.5.3)
# ============================================================

class JSONValidator:
    TYPES = {
        "per": (int, float), "croissance_ca": (int, float),
        "dette_nette_ebitda": (int, float), "roic": (int, float),
        "wacc": (int, float), "marge_nette": (int, float),
        "cash_conversion": (int, float), "div_yield": (int, float),
        "free_float": (int, float), "secteur": str,
        "catalyseur": str, "risque": str, "source": str, "date_maj": str,
    }

    def valider(self, fond):
        rapport = {}
        for ticker, data in fond.items():
            erreurs, warnings = [], []
            for c in CHAMPS_OBLIGATOIRES:
                if c not in data: erreurs.append(f"Champ obligatoire manquant: {c}")
            for c, types in self.TYPES.items():
                if c in data and data[c] is not None and not isinstance(data[c], types):
                    erreurs.append(f"Type incorrect pour {c}")
            if data.get("div_yield", 0) and data.get("div_yield", 0) > 25:
                warnings.append("Dividend yield suspect")
            if data.get("wacc") and not (5 <= data["wacc"] <= 20):
                warnings.append("WACC hors plage normale")
            rapport[ticker] = {
                "erreurs": erreurs, "warnings": warnings,
                "nb_erreurs": len(erreurs), "nb_warnings": len(warnings),
                "statut": "ERREUR" if erreurs else ("WARNING" if warnings else "OK")
            }
        nb_ok = sum(1 for r in rapport.values() if not r["nb_erreurs"] and not r["nb_warnings"])
        nb_err = sum(1 for r in rapport.values() if r["nb_erreurs"])
        logger.info(f"JSONValidator: {len(fond)} tickers | Erreurs: {nb_err} | OK: {nb_ok}")
        return rapport

# ============================================================
# MODULE SCORE QUALITE (v3.5.3)
# ============================================================

class FundamentalDataQualityScore:
    SCORING = {
        "per": 5, "croissance_ca": 5, "dette_nette_ebitda": 5, "roic": 5,
        "wacc": 5, "marge_nette": 5, "cash_conversion": 5, "div_yield": 5,
        "free_float": 7, "consensus_analystes": 8, "historique_previsionnel": 25,
        "tendance_roic": 5, "tendance_marge": 5, "tendance_dette": 5,
    }
    FIABILITE = {
        "attijari global research": 1.0, "agr": 1.0, "bkgr": 0.95,
        "bmce capital": 1.0, "cdgcapital": 0.95, "cfgbank": 0.90,
        "medias24": 0.85, "boursenews": 0.80, "bvc": 0.95, "ammc": 1.0,
    }

    def scorer(self, fond):
        scores = {}
        for ticker, data in fond.items():
            score = 0; max_score = sum(self.SCORING.values()); detail = {}
            for champ, pts in self.SCORING.items():
                if champ == "historique_previsionnel":
                    h = data.get(champ, {})
                    if h:
                        nb_a = len(h); nb_c = sum(len(v) for v in h.values())
                        obt = int(pts * min(nb_a/3, 1) * min(nb_c/12, 1))
                        score += obt; detail[champ] = f"{obt}/{pts}"
                    else: detail[champ] = f"0/{pts}"
                elif champ == "consensus_analystes":
                    ca = data.get(champ, [])
                    obt = pts if ca else 0; score += obt; detail[champ] = f"{obt}/{pts}"
                else:
                    obt = pts if (champ in data and data[champ] is not None) else 0
                    score += obt; detail[champ] = f"{obt}/{pts}"
            # Coefficient source
            src = data.get("source", "").lower()
            coef = max((v for k, v in self.FIABILITE.items() if k in src), default=0.9)
            # Fraicheur
            penalty = 0; fraicheur = "N/A"
            ds = data.get("date_maj")
            if ds:
                try:
                    delta = (datetime.today() - datetime.strptime(ds, "%Y-%m-%d")).days
                    fraicheur = f"{delta}j"
                    if delta > 90: penalty = 5
                    if delta > 180: penalty = 10
                except: pass
            score_final = max(0, round(score * coef) - penalty)
            grade = "A" if score_final >= 85 else "B" if score_final >= 70 else "C" if score_final >= 55 else "D" if score_final >= 40 else "F"
            scores[ticker] = {
                "score": score_final, "score_brut": score, "coef": coef,
                "grade": grade, "fraicheur": fraicheur, "detail": detail,
                "source": data.get("source", "")[:40],
            }
        for t, s in sorted(scores.items(), key=lambda x: -x[1]["score"])[:10]:
            logger.info(f"{t:<6} | {s['score']:>3}/100 | Grade {s['grade']} | {s['fraicheur']}")
        return scores


# ============================================================
# INDICATEURS TECHNIQUES (v3.4)
# ============================================================

class TechnicalIndicators:
    @staticmethod
    def calculate_all(df):
        df = df.copy()
        for p in [20, 50, 100, 200]:
            df[f"MA{p}"] = df["close"].rolling(p).mean()

        delta = df["close"].diff()
        gain = delta.clip(lower=0)
        loss = -delta.clip(upper=0)
        avg_gain = gain.ewm(alpha=1/14, adjust=False).mean()
        avg_loss = loss.ewm(alpha=1/14, adjust=False).mean()
        rs = avg_gain / avg_loss.replace(0, np.nan)
        df["RSI"] = (100 - 100/(1+rs)).fillna(50)

        ema12 = df["close"].ewm(span=12, adjust=False).mean()
        ema26 = df["close"].ewm(span=26, adjust=False).mean()
        df["MACD"] = ema12 - ema26
        df["MACD_SIGNAL"] = df["MACD"].ewm(span=9, adjust=False).mean()
        df["MACD_HIST"] = df["MACD"] - df["MACD_SIGNAL"]

        ma20 = df["close"].rolling(20).mean()
        std = df["close"].rolling(20).std()
        df["BB_upper"] = ma20 + 2*std
        df["BB_lower"] = ma20 - 2*std
        df["BB_width"] = (df["BB_upper"] - df["BB_lower"]) / ma20

        df["tenkan"] = (df["high"].rolling(9).max() + df["low"].rolling(9).min()) / 2
        df["kijun"]  = (df["high"].rolling(26).max() + df["low"].rolling(26).min()) / 2
        df["kumo_a"] = (df["tenkan"] + df["kijun"]) / 2
        df["kumo_b"] = (df["high"].rolling(52).max() + df["low"].rolling(52).min()) / 2

        ll = df["low"].rolling(14).min()
        hh = df["high"].rolling(14).max()
        df["Stoch_K"] = (100*(df["close"]-ll)/(hh-ll).replace(0,np.nan)).fillna(50)
        df["Stoch_D"] = df["Stoch_K"].rolling(3).mean().fillna(50)

        df["Vol_MA20"] = df["volume"].rolling(20).mean()
        df["Volume_Ratio"] = df["volume"] / df["Vol_MA20"]
        df["Return_1D"] = df["close"].pct_change() * 100
        df["Return_5D"] = df["close"].pct_change(5) * 100
        df["Return_20D"] = df["close"].pct_change(20) * 100
        df["Volatility_20"] = df["Return_1D"].rolling(20).std()

        tr = pd.concat([
            df["high"] - df["low"],
            abs(df["high"] - df["close"].shift(1)),
            abs(df["low"]  - df["close"].shift(1))
        ], axis=1).max(axis=1)
        df["ATR"] = tr.rolling(14).mean()
        df["High_20"] = df["high"].rolling(20).max()
        df["Low_20"]  = df["low"].rolling(20).min()
        return df

# ============================================================
# SCORING TECHNIQUE (v3.4)
# ============================================================

class TechnicalScorer:
    def score(self, df):
        m = SafeMath(); score = 0; maxs = 0; reasons = []
        p = m.last(df["close"]); ma20 = m.last(df["MA20"])
        ma50 = m.last(df["MA50"]); ma200 = m.last(df["MA200"])
        rsi = m.last(df["RSI"]); macd = m.last(df["MACD"])
        sig = m.last(df["MACD_SIGNAL"]); hist = m.last(df["MACD_HIST"])
        histp = m.prev(df["MACD_HIST"]); vr = m.last(df["Volume_Ratio"])
        sk = m.last(df["Stoch_K"]); sd = m.last(df["Stoch_D"])
        bbu = m.last(df["BB_upper"]); bbl = m.last(df["BB_lower"])
        bbw = m.last(df["BB_width"]); tenkan = m.last(df["tenkan"])
        kijun = m.last(df["kijun"]); ka = m.last(df["kumo_a"])
        kb = m.last(df["kumo_b"]); ret5 = m.last(df["Return_5D"])
        vol20 = m.last(df["Volatility_20"])

        nuage_h = max(ka,kb) if not pd.isna(ka) and not pd.isna(kb) else np.nan
        nuage_b = min(ka,kb) if not pd.isna(ka) and not pd.isna(kb) else np.nan

        maxs += 3
        if p > ma20 > ma50: score += 3; reasons.append("Tendance haussiere forte")
        elif p < ma20 < ma50: reasons.append("Tendance baissiere confirmee")
        else: score += 1; reasons.append("Moyennes mobiles mixtes")

        maxs += 2
        if not pd.isna(ma200):
            if p > ma200: score += 2 if p > ma200*1.05 else 1; reasons.append("Prix au-dessus MA200")
            else: reasons.append("Prix sous MA200")

        maxs += 3
        if macd > sig and hist > histp > 0: score += 3; reasons.append("MACD haussier accelerant")
        elif macd > sig: score += 2; reasons.append("MACD haussier")
        elif macd < sig and hist < histp: reasons.append("MACD baissier degrade")
        else: score += 0.5; reasons.append("MACD en stabilisation")

        maxs += 2
        # Memoire RSI — comparer avec RSI il y a 10 jours pour detecter accélération
        rsi_10j = SafeMath.last(df["RSI"].shift(10)) if "RSI" in df.columns else rsi
        rsi_trend = rsi - rsi_10j if not pd.isna(rsi_10j) else 0

        if 40 <= rsi <= 60: score += 2; reasons.append(f"RSI equilibre: {rsi:.1f}")
        elif 60 < rsi <= 70: score += 1.5; reasons.append(f"RSI positif a surveiller: {rsi:.1f}")
        elif rsi > 70 and rsi_trend > 15:
            score += 2; reasons.append(f"RSI {rsi:.1f} momentum confirme (+{rsi_trend:.0f}pts en 10j)")
        elif rsi > 70 and rsi_trend > 5:
            score += 1; reasons.append(f"RSI {rsi:.1f} accélération en cours")
        elif rsi > 70:
            reasons.append(f"RSI surchauffe sans momentum: {rsi:.1f}")
        else: score += 1; reasons.append(f"RSI bas: {rsi:.1f}")

        maxs += 3
        if not pd.isna(nuage_h):
            if p > nuage_h and tenkan > kijun: score += 3; reasons.append("Ichimoku positif")
            elif p > nuage_h: score += 2; reasons.append("Prix au-dessus du nuage")
            elif p > nuage_b: score += 1; reasons.append("Prix dans le nuage")
            else: reasons.append("Prix sous le nuage")

        maxs += 2
        if sk < 20 and sk > sd: score += 2; reasons.append("Stochastique sortie de survente")
        elif sk > 80: score += 0.5; reasons.append("Stochastique surachat")
        else: score += 1; reasons.append("Stochastique neutre")

        maxs += 2
        if not any(pd.isna(x) for x in [p,bbu,bbl]):
            if p < bbl and rsi < 40: score += 2; reasons.append("Rebond Bollinger possible")
            elif p > bbu and rsi > 65: reasons.append("Exces Bollinger")
            elif not pd.isna(bbw) and bbw < 0.05: score += 1.5; reasons.append("Squeeze Bollinger")
            else: score += 1; reasons.append("Bollinger neutre")

        maxs += 2
        if not pd.isna(vr):
            if vr >= 2 and p > ma20: score += 2; reasons.append("Volume fort confirme")
            elif vr >= 1.2: score += 1; reasons.append("Volume superieur moyenne")
            elif vr < 0.15: reasons.append("Volume tres faible")
            elif vr < 0.5: score += 0.5; reasons.append("Volume faible")
            else: score += 1; reasons.append("Volume modere")

        maxs += 1
        if not pd.isna(ret5):
            score += 1 if abs(ret5) <= 5 else 0.5
            reasons.append(f"Momentum 5j: {ret5:.1f}%")

        maxs += 1
        if not pd.isna(vol20):
            if vol20 < 1.5: score += 1; reasons.append(f"Volatilite faible: {vol20:.1f}%")
            elif vol20 <= 5: score += 0.5; reasons.append(f"Volatilite moderee: {vol20:.1f}%")
            else: reasons.append(f"Volatilite elevee: {vol20:.1f}%")

        # Fibonacci 52 semaines
        try:
            high_52w = df["high"].tail(252).max()
            low_52w  = df["low"].tail(252).min()
            diff     = high_52w - low_52w
            fib382   = round(high_52w - 0.382 * diff, 2)
            fib500   = round(high_52w - 0.500 * diff, 2)
            fib618   = round(high_52w - 0.618 * diff, 2)
            fib127   = round(high_52w + 0.127 * diff, 2)
            reasons.append(f"Fibo 38.2%={fib382} | 50%={fib500} | 61.8%={fib618} | R127%={fib127}")
        except:
            pass

        return ScoreResult(score, maxs, reasons, {
            "prix": p, "ma20": ma20, "ma50": ma50, "ma200": ma200,
            "rsi": rsi, "macd": macd, "vol_ratio": vr, "stoch_k": sk
        })

# ============================================================
# SCORING FONDAMENTAL (v3.4 + forward PER v4.0)
# ============================================================

class FundamentalScorer:
    def __init__(self, f):
        self.raw = f or {}
        self.math = SafeMath()
        self.f = self._normalize(self.raw)

    def _normalize(self, d):
        mapping = {
            "per": ["per","PER"], "forward_per": ["forward_per","FORWARD_PER"],
            "croissance_ca": ["croissance_ca","croissance","growth"],
            "dette_nette_ebitda": ["dette_nette_ebitda","dette"],
            "div_yield": ["div_yield","dividende"], "roic": ["roic","ROIC"],
            "wacc": ["wacc","WACC"], "cash_conversion": ["cash_conversion"],
            "marge_nette": ["marge_nette","marge"], "secteur": ["secteur"],
            "catalyseur": ["catalyseur"], "momentum_fondamental": ["momentum_fondamental"],
            "rerating": ["rerating"], "cycle_commodities": ["cycle_commodities"],
            "bpa_2026": ["bpa_2026","BPA_2026"], "dps_2026": ["dps_2026","DPS_2026"],
        }
        out = {}
        for k, alts in mapping.items():
            for a in alts:
                if a in d and d[a] is not None:
                    out[k] = d[a]; break
        return out

    def score(self):
        if not self.f:
            return ScoreResult(0, 10, ["Pas de donnees fondamentales"])
        m = self.math; score = 0; maxs = 0; r = []
        per = m.to_float(self.f.get("per"))
        fper = m.to_float(self.f.get("forward_per"))
        g = m.to_float(self.f.get("croissance_ca"))
        d = m.to_float(self.f.get("dette_nette_ebitda"))
        div = m.to_float(self.f.get("div_yield"))
        roic = m.to_float(self.f.get("roic"))
        wacc = m.to_float(self.f.get("wacc"))
        cash = m.to_float(self.f.get("cash_conversion"))
        marge = m.to_float(self.f.get("marge_nette"))

        maxs += 2
        if not pd.isna(per):
            if per < 8: score += 2; r.append("PER tres attractif")
            elif per < 12: score += 1.75; r.append("PER attractif")
            elif per < 18: score += 1.25; r.append("PER raisonnable")
            elif per < 25: score += 0.5; r.append("PER exigeant")
            else: r.append("PER eleve")

        # Forward PER bonus
        if not pd.isna(fper) and not pd.isna(per) and fper < per * 0.9:
            score += 0.5; r.append("Forward PER en rerating")

        maxs += 2
        if not pd.isna(g):
            if g > 20: score += 2; r.append("Croissance forte")
            elif g > 15: score += 1.5; r.append("Croissance solide")
            elif g > 8: score += 1; r.append("Croissance correcte")
            elif g > 0: score += 0.5; r.append("Croissance faible")
            else: r.append("Croissance negative")

        maxs += 2
        if not pd.isna(d):
            if d <= 1: score += 2; r.append("Dette tres faible")
            elif d <= 2: score += 1.75; r.append("Dette faible")
            elif d <= 3: score += 1; r.append("Dette moderee")
            elif d <= 4: score += 0.25; r.append("Dette elevee")
            else: r.append("Dette tres elevee")

        maxs += 2
        if not pd.isna(roic) and not pd.isna(wacc):
            s = roic - wacc
            if s >= 5: score += 2; r.append("Forte creation de valeur")
            elif s >= 2: score += 1.5; r.append("Creation de valeur")
            elif s >= 0: score += 0.5; r.append("Creation de valeur faible")
            else: r.append("ROIC inferieur au WACC")

        maxs += 1
        if not pd.isna(cash):
            if cash >= 80: score += 1; r.append("Cash conversion excellente")
            elif cash >= 65: score += 0.75; r.append("Cash conversion correcte")
            elif cash >= 50: score += 0.5; r.append("Cash conversion moyenne")
            else: r.append("Cash conversion faible")

        maxs += 1
        if not pd.isna(marge):
            if marge >= 18: score += 1; r.append("Marge nette elevee")
            elif marge >= 12: score += 0.75; r.append("Marge solide")
            elif marge >= 7: score += 0.5; r.append("Marge correcte")
            else: r.append("Marge faible")

        maxs += 1
        if not pd.isna(div):
            if div > 4: score += 1; r.append("Dividende eleve")
            elif div > 2.5: score += 0.75; r.append("Dividende correct")
            elif div > 1: score += 0.25; r.append("Dividende modeste")

        return ScoreResult(score, maxs, r)


# ============================================================
# RED FLAGS (v3.4)
# ============================================================

class RedFlagDetector:
    def __init__(self, f):
        self.f = FundamentalScorer(f).f
        self.m = SafeMath()

    def detect(self, df):
        flags = []
        d = self.m.to_float(self.f.get("dette_nette_ebitda"))
        cash = self.m.to_float(self.f.get("cash_conversion"))
        roic = self.m.to_float(self.f.get("roic"))
        wacc = self.m.to_float(self.f.get("wacc"))
        marge = self.m.to_float(self.f.get("marge_nette"))
        p = self.m.last(df["close"]); ma20 = self.m.last(df["MA20"])
        ma50 = self.m.last(df["MA50"]); rsi = self.m.last(df["RSI"])
        vr = self.m.last(df["Volume_Ratio"]); ret20 = self.m.last(df["Return_20D"])

        if not pd.isna(d):
            if d > 5: flags.append(RedFlag(Severity.CRITIQUE, "Structure", "Dette/EBITDA critique", True))
            elif d > 4: flags.append(RedFlag(Severity.ELEVEE, "Structure", "Dette elevee"))
            elif d > 3: flags.append(RedFlag(Severity.MOYENNE, "Structure", "Endettement eleve"))

        if not pd.isna(cash):
            if cash < 30: flags.append(RedFlag(Severity.CRITIQUE, "Cash-flow", "Cash conversion critique", True))
            elif cash < 50: flags.append(RedFlag(Severity.MOYENNE, "Cash-flow", "Cash conversion faible"))

        if not pd.isna(roic) and not pd.isna(wacc) and roic < wacc:
            flags.append(RedFlag(Severity.ELEVEE, "Rentabilite", "ROIC inferieur au WACC"))

        if not pd.isna(marge) and marge < 5:
            flags.append(RedFlag(Severity.MOYENNE, "Rentabilite", "Marge nette faible"))

        if p < ma20 < ma50:
            flags.append(RedFlag(Severity.ELEVEE, "Technique", "Tendance baissiere court/moyen terme"))

        if not pd.isna(rsi):
            if rsi > 80: flags.append(RedFlag(Severity.CRITIQUE, "Technique", "RSI surchauffe extreme", True))
            elif rsi > 75: flags.append(RedFlag(Severity.ELEVEE, "Technique", "RSI surchauffe"))

        if not pd.isna(vr):
            if vr < 0.15: flags.append(RedFlag(Severity.CRITIQUE, "Liquidite", "Volume tres faible", True))
            elif vr < 0.3: flags.append(RedFlag(Severity.MOYENNE, "Liquidite", "Volume faible"))

        if not pd.isna(ret20):
            if ret20 < -25: flags.append(RedFlag(Severity.CRITIQUE, "Momentum", "Correction forte sur 20j", True))
            elif ret20 < -15: flags.append(RedFlag(Severity.ELEVEE, "Momentum", "Correction sur 20j"))

        return flags

# ============================================================
# SETUP DETECTOR (v3.4)
# ============================================================

class SetupDetector:
    def detect(self, df):
        m = SafeMath()
        p = m.last(df["close"]); ma20 = m.last(df["MA20"])
        ma50 = m.last(df["MA50"]); rsi = m.last(df["RSI"])
        vr = m.last(df["Volume_Ratio"])
        high20 = m.last(df["High_20"]); low20 = m.last(df["Low_20"])
        macd = m.last(df["MACD"]); sig = m.last(df["MACD_SIGNAL"])

        has_trend = not any(pd.isna(x) for x in [p, ma20, ma50])
        has_rsi   = not pd.isna(rsi)
        has_macd  = not any(pd.isna(x) for x in [macd, sig])
        has_vr    = not pd.isna(vr)

        if has_trend and has_rsi and p < ma20 < ma50 and rsi < 48: return Setup.FAIBLESSE
        if has_trend and has_rsi and has_macd and p > ma20 > ma50 and rsi > 52 and macd > sig:
            if not has_vr or vr >= 1.2: return Setup.MOMENTUM_CONFIRME
        if has_macd and not pd.isna(high20) and p >= high20*0.98 and macd > sig:
            if not has_vr or vr >= 1.3: return Setup.BREAKOUT_POTENTIEL
        if has_trend and has_rsi and p > ma50 and ma20 > ma50 and 40 <= rsi <= 58 and p <= ma20*1.02: return Setup.PULLBACK_HAUSSIER
        if has_rsi and not pd.isna(low20) and (rsi < 38 or (not pd.isna(m.last(df["Stoch_K"])) and m.last(df["Stoch_K"]) < 25)) and p <= low20*1.06: return Setup.CONTRARIEN
        if has_trend and has_macd and p > ma50 and has_vr and vr < 0.8 and has_rsi and 45 <= rsi <= 60 and macd > sig: return Setup.ACCUMULATION_LENTE
        return Setup.NEUTRE

# ============================================================
# PONDERATION DYNAMIQUE LIQUIDITE (v4.0)
# ============================================================

class LiquidityAnalyzer:
    TIERS = [
        (5_000_000, 0.60, 0.40),   # Tier 1: Large Cap
        (1_000_000, 0.55, 0.45),   # Tier 2: Mid Cap
        (200_000,   0.50, 0.50),   # Tier 3: Small Cap
        (50_000,    0.40, 0.60),   # Tier 4: Micro Cap
        (0,         0.30, 0.70),   # Tier 5: Illiquide
    ]

    @classmethod
    def get_weights(cls, vol_moyen_20j: float) -> Tuple[float, float]:
        for seuil, w_tech, w_fond in cls.TIERS:
            if vol_moyen_20j >= seuil:
                return w_tech, w_fond
        return 0.30, 0.70

    @classmethod
    def get_discount(cls, volume_ratio: float) -> float:
        """Decote valorisation 0-20% selon liquidite"""
        if volume_ratio >= 1.5: return 0.0
        if volume_ratio >= 1.0: return 0.05
        if volume_ratio >= 0.5: return 0.10
        if volume_ratio >= 0.3: return 0.15
        return 0.20

# ============================================================
# EVA CALCULATOR (v4.0)
# ============================================================

class EVACalculator:
    @staticmethod
    def calculate(roic, wacc):
        if pd.isna(roic) or pd.isna(wacc): return None, "Donnees insuffisantes"
        spread = roic - wacc
        if spread >= 5: return spread, "Creation de valeur forte"
        if spread >= 2: return spread, "Creation de valeur"
        if spread >= 0: return spread, "Creation de valeur faible"
        return spread, "Destruction de valeur"

# ============================================================
# DECISION ENGINE (v3.4 + ACHAT FORT v4.0)
# ============================================================

class DecisionEngine:
    def decide(self, st, sf, setup, flags, vol_moyen_20j=0, volume_ratio=1.0, fond=None, df=None):
        blocking = [f for f in flags if f.is_blocking]
        w_tech, w_fond = LiquidityAnalyzer.get_weights(vol_moyen_20j)
        score = round(st.normalized * w_tech + sf.normalized * w_fond, 2)

        # BONUS SECTORIEL — Mines et commodites en cycle bullish
        bonus_sectoriel = 0.0
        if fond:
            secteur  = str(fond.get("secteur", "")).lower()
            cycle    = str(fond.get("cycle", "")).lower()
            catalys  = str(fond.get("catalyseur", "")).lower()
            mots_mines = ["mine", "argent", "or", "cuivre", "zinc", "plomb", "cobalt", "metal"]
            is_mine  = any(s in secteur for s in mots_mines)
            is_bull  = cycle == "bullish" or any(s in catalys for s in ["or", "argent", "cuivre", "metal"])
            if is_mine and is_bull:
                bonus_sectoriel = 1.0
            elif is_mine:
                bonus_sectoriel = 0.5
            score = round(min(score + bonus_sectoriel, 10.0), 2)

        # FIBONACCI — supports et resistances depuis historique 52 semaines
        fib_info = ""
        if df is not None and len(df) >= 50:
            try:
                high_52w = df["high"].tail(252).max()
                low_52w  = df["low"].tail(252).min()
                prix     = SafeMath.last(df["close"])
                diff     = high_52w - low_52w
                fib_618  = round(high_52w - 0.618 * diff, 2)
                fib_500  = round(high_52w - 0.500 * diff, 2)
                fib_382  = round(high_52w - 0.382 * diff, 2)
                fib_127  = round(high_52w + 0.127 * diff, 2)
                # Bonus si prix proche d un support Fibonacci (+/- 3%)
                for fib, nom in [(fib_618, "61.8%"), (fib_500, "50%"), (fib_382, "38.2%")]:
                    if abs(prix - fib) / fib < 0.03:
                        score = round(min(score + 0.5, 10.0), 2)
                        fib_info = f"Support Fibo {nom} a {fib} DH"
                        break
                if not fib_info:
                    fib_info = f"Fibo 38.2%={fib_382} | 50%={fib_500} | 61.8%={fib_618}"
            except:
                fib_info = ""

        # 1 CRITIQUE = EVITER immediat
        if len(blocking) >= 1:
            return Signal.EVITER, f"EVITER — risque bloquant: {blocking[0].message}", score

        if score >= 8.0 and setup in [Setup.MOMENTUM_CONFIRME, Setup.BREAKOUT_POTENTIEL] and len(blocking) == 0:
            return Signal.ACHAT_FORT, f"ACHAT FORT — {setup.value}", score
        if score >= 7 and setup in [Setup.MOMENTUM_CONFIRME, Setup.BREAKOUT_POTENTIEL]:
            return Signal.ACHETER, f"ACHETER — {setup.value}", score
        if score >= 6.5 and setup == Setup.PULLBACK_HAUSSIER:
            return Signal.ACHETER, "ACHETER — pullback haussier", score
        if score >= 6.5 and bonus_sectoriel > 0:
            return Signal.ACHETER, f"ACHETER — cycle commodites {fib_info}", score
        if score >= 6 and setup == Setup.ACCUMULATION_LENTE:
            return Signal.ACCUMULER, "ACCUMULER — accumulation progressive", score
        if score >= 5.5:
            return Signal.SURVEILLER, f"SURVEILLER — {setup.value}", score
        if score >= 4:
            return Signal.ATTENDRE, f"ATTENDRE — {setup.value}", score
        return Signal.EVITER, f"EVITER — {setup.value}", score

# ============================================================
# VALORISATION (v3.4 + targets JSON v4.0)
# ============================================================

class ValuationEngine:
    def __init__(self, f):
        self.f = FundamentalScorer(f).f
        self.m = SafeMath()

    def calculate(self, price, volume_ratio=1.0):
        # Priorite 1: targets JSON
        target = self.m.to_float(self.f.get("target_price"))
        bear_t = self.m.to_float(self.f.get("bear_case"))
        bull_t = self.m.to_float(self.f.get("bull_case"))

        if not pd.isna(target) and target > 0:
            discount = LiquidityAnalyzer.get_discount(volume_ratio)
            base = round(target * (1 - discount), 2)
            bull = round(bull_t * (1 - discount), 2) if not pd.isna(bull_t) else round(base * 1.25, 2)
            bear = round(bear_t * (1 - discount), 2) if not pd.isna(bear_t) else round(base * 0.78, 2)
            return Valuation(
                base, bull, bear,
                self.m.pct_change(base, price), self.m.pct_change(bull, price),
                self.m.pct_change(bear, price), round(discount * 100, 1),
                "Targets analystes (ajuste liquidite)", "Haute" if discount == 0 else "Moyenne"
            )

        # Priorite 2: calcul multi-facteurs
        per = self.m.to_float(self.f.get("per"))
        g = self.m.to_float(self.f.get("croissance_ca"))
        roic = self.m.to_float(self.f.get("roic"))
        wacc = self.m.to_float(self.f.get("wacc"))
        d = self.m.to_float(self.f.get("dette_nette_ebitda"))
        cash = self.m.to_float(self.f.get("cash_conversion"))

        if pd.isna(per) or pd.isna(g):
            return Valuation(None, None, None, None, None, None, None, "Insuffisant", "Faible")

        prem = 0
        if g > 20: prem += .15
        elif g > 15: prem += .10
        elif g > 8: prem += .05
        elif g < 0: prem -= .10

        if not pd.isna(roic) and not pd.isna(wacc):
            s = roic - wacc
            if s >= 5: prem += .12
            elif s >= 2: prem += .07
            elif s < 0: prem -= .10

        if per < 8: prem += .12
        elif per < 12: prem += .08
        elif per > 28: prem -= .15
        elif per > 22: prem -= .08

        if not pd.isna(cash):
            if cash >= 80: prem += .05
            elif cash < 40: prem -= .08

        if not pd.isna(d):
            if d > 4: prem -= .10
            elif d > 3: prem -= .05
            elif d < 1: prem += .05

        # Decote liquidite
        discount = LiquidityAnalyzer.get_discount(volume_ratio)
        prem -= discount

        base = round(price * (1 + prem), 2)
        bull = round(base * 1.25, 2)
        bear = round(base * 0.78, 2)
        return Valuation(
            base, bull, bear,
            self.m.pct_change(base, price), self.m.pct_change(bull, price),
            self.m.pct_change(bear, price), round(prem * 100, 1),
            "Multi-facteurs (ajuste liquidite)", "Moyenne" if abs(prem) < 0.2 else "Haute"
        )


# ============================================================
# NOTE INSTITUTIONNELLE (v3.4 enrichie 8 sections)
# ============================================================

class NoteGenerator:
    _SEV_ORDER = {Severity.CRITIQUE: 0, Severity.ELEVEE: 1, Severity.MOYENNE: 2, Severity.FAIBLE: 3}

    def generate(self, r):
        f = FundamentalScorer(r.fundamental_data).f
        live = r.live_data or {}
        tech = "\n".join([f"  * {x}" for x in r.score_technical.reasons])
        fond = "\n".join([f"  * {x}" for x in r.score_fundamental.reasons])

        # Red flags tries par severite
        sorted_flags = sorted(r.red_flags, key=lambda x: self._SEV_ORDER.get(x.severity, 99))
        flags = "\n".join([
            f"  {x.severity.value} | {x.category} : {x.message}" + (" [BLOQUANT]" if x.is_blocking else "")
            for x in sorted_flags
        ]) if r.red_flags else "  OK Aucun red flag majeur"

        # EVA
        roic = SafeMath.to_float(f.get("roic"))
        wacc = SafeMath.to_float(f.get("wacc"))
        eva_spread, eva_verdict = EVACalculator.calculate(roic, wacc)
        eva_str = f"{eva_spread:.1f}pts — {eva_verdict}" if eva_spread is not None else "N/A"

        # Pondération
        vol_moy = live.get("volume_mad", 0)
        w_tech, w_fond = LiquidityAnalyzer.get_weights(vol_moy)

        return f"""
================================================================================
NOTE INSTITUTIONNELLE — {r.ticker}
BVC ANALYZER v5.0 | {r.date.strftime('%d/%m/%Y %H:%M')} | Africa/Casablanca
================================================================================

1. SYNTHESE EXECUTIVE
   Prix        : {r.price:.2f} DH
   Signal      : {r.action}
   Setup       : {r.setup.value}
   Score Tech  : {r.score_technical.normalized}/10 (poids {w_tech:.0%})
   Score Fond  : {r.score_fundamental.normalized}/10 (poids {w_fond:.0%})
   Score Global: {r.score_global}/10
   Source      : {live.get('source','Medias24')}
   Cours live  : {live.get('cours','N/A')} DH
   Variation   : {live.get('variation','N/A')} %
   Volume      : {live.get('volume_mad','N/A')} DH
   Marche      : {r.market_status}

2. ANALYSE TECHNIQUE
{tech}

3. ANALYSE FONDAMENTALE
   PER       : {f.get('per','N/A')}x | Forward PER: {f.get('forward_per','N/A')}x
   Croissance: {f.get('croissance_ca','N/A')}% | Dette: {f.get('dette_nette_ebitda','N/A')}x
   ROIC/WACC : {f.get('roic','N/A')}/{f.get('wacc','N/A')} | EVA: {eva_str}
   Cash conv.: {f.get('cash_conversion','N/A')}% | Marge: {f.get('marge_nette','N/A')}%
   Dividende : {f.get('div_yield','N/A')}% | BPA 2026: {f.get('bpa_2026','N/A')} | DPS 2026: {f.get('dps_2026','N/A')}
{fond}

4. RED FLAGS (tries par severite)
{flags}

5. VALORISATION
   Methodologie: {r.valuation.methodology}
   Confiance   : {r.valuation.confidence}
   Premium     : {r.valuation.premium_pct}%
   Bear        : {r.valuation.bear} DH ({r.valuation.downside}%)
   Base        : {r.valuation.base} DH (+{r.valuation.upside_base}%)
   Bull        : {r.valuation.bull} DH (+{r.valuation.upside_bull}%)

6. EVA (Economic Value Added)
   {eva_str}

7. CATALYSEURS ET RISQUES
   Catalyseur : {f.get('catalyseur','N/A')}
   Risque     : {f.get('risque','N/A')}
   Secteur    : {f.get('secteur','N/A')}
   Momentum   : {f.get('momentum_fondamental','N/A')}
   Rerating   : {f.get('rerating','N/A')}
   Cycle      : {f.get('cycle_commodities','N/A')}

8. CONCLUSION
   Recommandation : {r.action}
   Lecture        : Technique + Fondamental + Liquidite + EVA + Red Flags
   Methodologie   : BVC Analyzer v5.2 — RSI Memoire + Fibonacci + Bonus Mines
================================================================================
"""

# ============================================================
# EXPORT ENGINE (v3.5.3 enrichi)
# ============================================================

class ExportEngine:
    def _hdr(self, ws, row, ncols, bg="1F4E79"):
        fill = PatternFill("solid", start_color=bg)
        font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        aln = Alignment(horizontal="center", vertical="center")
        for c in range(1, ncols+1):
            cell = ws.cell(row, c)
            cell.fill = fill; cell.font = font; cell.alignment = aln

    def _borders(self, ws, r1, r2, c1, c2):
        t = Side(style="thin")
        b = Border(left=t, right=t, top=t, bottom=t)
        for row in ws.iter_rows(r1, r2, c1, c2):
            for cell in row: cell.border = b

    def _col_widths(self, ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def rapport_data_quality(self, scores, fond, path=None):
        path = path or os.path.join(OUTPUT_DIR, "rapport_data_quality.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Scores Qualite"
        h1 = ["Ticker", "Score /100", "Grade", "Fraicheur", "Coef Source",
              "Champs Base", "Historique", "Consensus", "Tendances", "Source"]
        for j, h in enumerate(h1, 1): ws1.cell(1, j, h)
        self._hdr(ws1, 1, len(h1))

        GRADE_COLOR = {"A": "C6EFCE", "B": "FFEB9C", "C": "FFCC99", "D": "FFC7CE", "F": "FF0000"}
        for i, (ticker, s) in enumerate(sorted(scores.items(), key=lambda x: -x[1]["score"]), 2):
            d = fond.get(ticker, {})
            histo = d.get("historique_previsionnel", {})
            nb_b = sum(1 for c in CHAMPS_OBLIGATOIRES if c in d)
            ws1.cell(i, 1, ticker)
            ws1.cell(i, 2, s["score"])
            ws1.cell(i, 3, s["grade"])
            ws1.cell(i, 4, s["fraicheur"])
            ws1.cell(i, 5, f"{s['coef']:.0%}")
            ws1.cell(i, 6, f"{nb_b}/{len(CHAMPS_OBLIGATOIRES)}")
            ws1.cell(i, 7, f"{len(histo)}ans" if histo else "—")
            ws1.cell(i, 8, "✓" if d.get("consensus_analystes") else "—")
            ws1.cell(i, 9, "✓" if d.get("tendance_roic") else "—")
            ws1.cell(i, 10, s["source"])
            c = GRADE_COLOR.get(s["grade"], "FFFFFF")
            ws1.cell(i, 3).fill = PatternFill("solid", start_color=c)
        self._col_widths(ws1, [10, 11, 8, 10, 12, 13, 12, 11, 11, 35])
        self._borders(ws1, 1, len(scores)+1, 1, len(h1))

        # Feuille 2 — Detail champs
        ws2 = wb.create_sheet("Detail Champs")
        extras = ["free_float", "forward_per", "bpa_2026", "dps_2026",
                  "tendance_roic", "tendance_marge", "tendance_dette",
                  "consensus_analystes", "historique_previsionnel"]
        h2 = ["Ticker"] + CHAMPS_OBLIGATOIRES + extras
        for j, h in enumerate(h2, 1): ws2.cell(1, j, h)
        self._hdr(ws2, 1, len(h2), "375623")
        for i, (ticker, d) in enumerate(fond.items(), 2):
            ws2.cell(i, 1, ticker)
            for j, c in enumerate(h2[1:], 2):
                val = d.get(c)
                if isinstance(val, (dict, list)): txt = f"({len(val)})"
                elif val is None: txt = "—"
                else: txt = str(val)[:30]
                ws2.cell(i, j, txt)
                if val is None: ws2.cell(i, j).fill = PatternFill("solid", start_color="FFE0E0")
        self._col_widths(ws2, [10] + [10]*len(h2[1:]))
        self._borders(ws2, 1, len(fond)+1, 1, len(h2))

        wb.save(path)
        logger.info(f"rapport_data_quality.xlsx -> {path}")
        return path

    def rapport_sources_pdf(self, resultats_pdf, log_traces, path=None):
        path = path or os.path.join(OUTPUT_DIR, "rapport_sources_pdf.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Resume PDFs"
        h1 = ["Ticker", "Fichier PDF", "Statut", "Cours Cible", "Reco", "Upside%",
              "Free Float%", "Analyste", "Annees", "Champs", "URL"]
        for j, h in enumerate(h1, 1): ws1.cell(1, j, h)
        self._hdr(ws1, 1, len(h1), "375623")

        for i, (ticker, d) in enumerate(resultats_pdf.items(), 2):
            meta = d.get("meta", {})
            histo = d.get("histo", {})
            nb_c = sum(len(v) for v in histo.values())
            ok = not d.get("erreur")
            ws1.cell(i, 1, ticker)
            ws1.cell(i, 2, d.get("url", "").split("/")[-1][:45])
            ws1.cell(i, 3, "OK" if ok else f"{d.get('erreur','')[:30]}")
            ws1.cell(i, 4, meta.get("cours_cible", "—"))
            ws1.cell(i, 5, meta.get("recommandation", "—"))
            ws1.cell(i, 6, meta.get("upside_pct", "—"))
            ws1.cell(i, 7, meta.get("free_float", "—"))
            ws1.cell(i, 8, meta.get("analyste", "—"))
            ws1.cell(i, 9, len(histo))
            ws1.cell(i, 10, nb_c)
            ws1.cell(i, 11, d.get("url", "")[:80])
            if not ok: ws1.cell(i, 3).fill = PatternFill("solid", start_color="FFC7CE")
        self._col_widths(ws1, [8, 48, 10, 13, 12, 9, 12, 25, 8, 8, 60])
        self._borders(ws1, 1, len(resultats_pdf)+1, 1, len(h1))

        # Feuille 2 — Log tracabilite
        ws2 = wb.create_sheet("Log Tracabilite")
        h2 = ["Ticker", "Source PDF", "Page", "Methode", "Annee", "Champ", "Valeur", "Confiance", "Ligne source"]
        for j, h in enumerate(h2, 1): ws2.cell(1, j, h)
        self._hdr(ws2, 1, len(h2), "5C3317")
        for i, trace in enumerate(log_traces, 2):
            ws2.cell(i, 1, trace.get("ticker", ""))
            ws2.cell(i, 2, trace.get("source", ""))
            ws2.cell(i, 3, trace.get("page", ""))
            ws2.cell(i, 4, trace.get("methode", ""))
            ws2.cell(i, 5, trace.get("annee", ""))
            ws2.cell(i, 6, trace.get("champ", ""))
            ws2.cell(i, 7, trace.get("valeur", ""))
            ws2.cell(i, 8, trace.get("confiance", ""))
            ws2.cell(i, 9, trace.get("ligne", "")[:80])
            if trace.get("confiance") == "moyenne":
                ws2.cell(i, 8).fill = PatternFill("solid", start_color="FFEB9C")
        self._col_widths(ws2, [8, 45, 6, 10, 8, 20, 10, 10, 80])
        if len(log_traces) > 0: ws2.freeze_panes = "A2"

        wb.save(path)
        logger.info(f"rapport_sources_pdf.xlsx -> {path} ({len(log_traces)} traces)")
        return path

    def rapport_decision(self, results, path=None):
        path = path or os.path.join(OUTPUT_DIR, f"rapport_decision_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Decisions"
        h = ["Ticker", "Prix", "Signal", "Action", "Score Global", "Technique", "Fondamental",
             "Setup", "Red Flags", "Base", "Upside%", "Marche", "Heure", "Source", "Cours Live", "Variation"]
        for j, col in enumerate(h, 1): ws.cell(1, j, col)
        self._hdr(ws, 1, len(h), "1F4E79")

        SIGNAL_COLOR = {
            Signal.ACHAT_FORT: "C6EFCE", Signal.ACHETER: "C6EFCE",
            Signal.ACCUMULER: "FFEB9C", Signal.SURVEILLER: "FFCC99",
            Signal.ATTENDRE: "FFC7CE", Signal.EVITER: "FF0000"
        }

        for i, r in enumerate(sorted(results, key=lambda x: x.score_global, reverse=True), 2):
            live = r.live_data or {}
            ws.cell(i, 1, r.ticker)
            ws.cell(i, 2, r.price)
            ws.cell(i, 3, r.signal.value)
            ws.cell(i, 4, r.action)
            ws.cell(i, 5, r.score_global)
            ws.cell(i, 6, r.score_technical.normalized)
            ws.cell(i, 7, r.score_fundamental.normalized)
            ws.cell(i, 8, r.setup.value)
            ws.cell(i, 9, len(r.red_flags))
            ws.cell(i, 10, r.valuation.base)
            ws.cell(i, 11, r.valuation.upside_base)
            ws.cell(i, 12, r.market_status)
            ws.cell(i, 13, r.date.strftime("%H:%M"))
            ws.cell(i, 14, live.get("source", "Medias24"))
            ws.cell(i, 15, live.get("cours", "N/A"))
            ws.cell(i, 16, live.get("variation", "N/A"))
            c = SIGNAL_COLOR.get(r.signal, "FFFFFF")
            ws.cell(i, 3).fill = PatternFill("solid", start_color=c)

        self._col_widths(ws, [10, 10, 14, 45, 12, 12, 12, 20, 12, 12, 10, 12, 8, 12, 12, 10])
        self._borders(ws, 1, len(results)+1, 1, len(h))
        wb.save(path)
        logger.info(f"rapport_decision.xlsx -> {path}")
        return path

    def export_json(self, fond, path=None):
        path = path or os.path.join(OUTPUT_DIR, "fondamentaux_enrichi.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(fond, f, ensure_ascii=False, indent=2)
        logger.info(f"fondamentaux_enrichi.json -> {path}")
        return path

# ============================================================
# BACKTESTER BVC (v3.5.3)
# ============================================================

class BacktesterBVC:
    def __init__(self, capital=100_000, buy_fee=BUY_FEE, sell_fee=SELL_FEE,
                 tax_rate=TAX_RATE, atr_mult=ATR_MULT, min_vol_dh=MIN_VOL_DH, split_excl=SPLIT_EXCL):
        self.capital = capital
        self.buy_fee = buy_fee
        self.sell_fee = sell_fee
        self.tax_rate = tax_rate
        self.atr_mult = atr_mult
        self.min_vol_dh = min_vol_dh
        self.split_excl = split_excl

    def _atr(self, df, n=14):
        if {"high", "low"}.issubset(df.columns):
            tr = df["high"] - df["low"]
        else:
            tr = df["close"].diff().abs().fillna(0)
        return tr.rolling(n).mean()

    def _cout_achat(self, montant): return montant * self.buy_fee
    def _cout_vente(self, montant): return montant * self.sell_fee
    def _impot_pv(self, pv): return max(0.0, pv) * self.tax_rate

    def run(self, df, ticker, signaux, split_dates=None):
        df = df.copy().reset_index(drop=True)
        df["date"] = pd.to_datetime(df["date"])
        df["atr"] = self._atr(df)
        if "volume" in df.columns:
            df["vol_dh"] = df["close"] * df["volume"]

        # Marquer post-split
        df["exclu"] = False
        if split_dates:
            for sd in split_dates:
                sd = pd.to_datetime(sd)
                mask = (df["date"] >= sd) & (df["date"] <= sd + pd.Timedelta(days=self.split_excl))
                df.loc[mask, "exclu"] = True

        capital = float(self.capital)
        position = 0
        prix_entree = 0.0
        stop = 0.0
        trades = []
        courbe = []

        for i, row in df.iterrows():
            cours = float(row["close"])
            val_ptf = capital + position * cours
            courbe.append(val_ptf)

            if row.get("exclu", False): continue

            # Filtre liquidite
            vol_ok = True
            if "vol_dh" in df.columns:
                vol_ok = float(row.get("vol_dh", self.min_vol_dh)) >= self.min_vol_dh

            # Stop ATR
            if position > 0 and cours < stop:
                montant = position * cours
                frais = self._cout_vente(montant)
                pv = (cours - prix_entree) * position
                impot = self._impot_pv(pv)
                capital += montant - frais - impot
                trades.append({
                    "type": "VENTE_STOP", "date": str(row["date"].date()),
                    "prix": cours, "qte": position,
                    "pv": round(pv, 2), "frais": round(frais, 2), "impot": round(impot, 2),
                })
                position = 0
                continue

            sig = bool(signaux.iloc[i]) if i < len(signaux) else False

            # Achat
            if sig and position == 0 and vol_ok:
                budget = capital * 0.95
                frais = self._cout_achat(budget)
                net = budget - frais
                qte = int(net / cours)
                if qte > 0:
                    prix_entree = cours
                    stop = prix_entree - self.atr_mult * float(row.get("atr", 0))
                    capital -= qte * cours + frais
                    position = qte
                    trades.append({
                        "type": "ACHAT", "date": str(row["date"].date()),
                        "prix": cours, "qte": qte, "stop": round(stop, 2),
                        "frais": round(frais, 2),
                    })

            # Vente signal
            elif not sig and position > 0:
                montant = position * cours
                frais = self._cout_vente(montant)
                pv = (cours - prix_entree) * position
                impot = self._impot_pv(pv)
                capital += montant - frais - impot
                trades.append({
                    "type": "VENTE_SIGNAL", "date": str(row["date"].date()),
                    "prix": cours, "qte": position,
                    "pv": round(pv, 2), "frais": round(frais, 2), "impot": round(impot, 2),
                })
                position = 0

        # Valeur finale
        val_finale = capital + position * float(df.iloc[-1]["close"])

        # Buy & Hold avec frais reels
        p0 = float(df.iloc[0]["close"])
        p1 = float(df.iloc[-1]["close"])
        qte_bh = int(self.capital * 0.95 / p0)
        cout_bh = self._cout_achat(qte_bh * p0) + self._cout_vente(qte_bh * p1)
        pv_bh = (p1 - p0) * qte_bh
        impot_bh = self._impot_pv(pv_bh)
        bh_fin = self.capital - qte_bh * p0 - self._cout_achat(qte_bh * p0) + qte_bh * p1 - self._cout_vente(qte_bh * p1) - impot_bh

        rend_s = (val_finale / self.capital - 1) * 100
        rend_bh = (bh_fin / self.capital - 1) * 100
        nb_t = len([t for t in trades if t["type"] == "ACHAT"])
        wins = len([t for t in trades if t.get("pv", 0) > 0])
        wr = wins / nb_t * 100 if nb_t else 0

        res = {
            "ticker": ticker, "capital_initial": self.capital,
            "val_finale": round(val_finale, 2),
            "rend_strategie": round(rend_s, 2),
            "rend_buy_hold": round(rend_bh, 2),
            "alpha": round(rend_s - rend_bh, 2),
            "nb_trades": nb_t, "win_rate": round(wr, 1),
            "frais_total": round(sum(t.get("frais", 0) for t in trades), 2),
            "impots_total": round(sum(t.get("impot", 0) for t in trades), 2),
            "courbe_capital": courbe, "trades": trades,
        }
        logger.info(f"Backtest {ticker}: Strategie {rend_s:+.1f}% | B&H {rend_bh:+.1f}% | Alpha {res['alpha']:+.1f}%")
        return res


# ============================================================
# ANALYSEUR PRINCIPAL (Orchestration v3.4 + pipeline v3.5.3)
# ============================================================

class BVCAnalyzer:
    def __init__(self):
        self.connector = BVCLiveConnector(ISIN_MAP)
        self.history = HistoryStore()
        self.fundamentals_cache = {}
        self.last_fetch = None
        self.pdf_extractor = None
        self.validator = JSONValidator()
        self.quality_scorer = FundamentalDataQualityScore()
        self.export_engine = ExportEngine()

    def fetch_fundamentals(self):
        if self.last_fetch and datetime.now() - self.last_fetch < timedelta(hours=6):
            return
        try:
            r = requests.get(GITHUB_URL, timeout=20)
            r.raise_for_status()
            self.fundamentals_cache = r.json()
            self.last_fetch = datetime.now()
            logger.info(f"Fondamentaux GitHub: {len(self.fundamentals_cache)} tickers")
        except Exception as e:
            logger.warning(f"GitHub indisponible: {e}")

    def run_pdf_pipeline(self):
        """Execute le pipeline d'enrichissement PDF"""
        logger.info("=" * 60)
        logger.info("PIPELINE ENRICHISSEMENT PDF v5.0")
        logger.info("=" * 60)

        fond_path = ensure_fondamentaux_json()
        self.pdf_extractor = FundamentalPDFExtractor(delay=2.5)
        resultats_pdf = self.pdf_extractor.extraire_tous()
        fond, path_json = self.pdf_extractor.enrichir_json(fond_path)

        # Validation
        rapport = self.validator.valider(fond)

        # Score qualite
        scores = self.quality_scorer.scorer(fond)

        # Exports
        path_dq = self.export_engine.rapport_data_quality(scores, fond)
        path_pdf = self.export_engine.rapport_sources_pdf(resultats_pdf, self.pdf_extractor.log)

        # Push GitHub
        if GITHUB_TOKEN:
            _push_github(path_json, GITHUB_TOKEN)

        logger.info("=" * 60)
        logger.info("PIPELINE PDF TERMINE")
        logger.info(f"JSON enrichi: {path_json}")
        logger.info(f"Rapport qualite: {path_dq}")
        logger.info(f"Rapport PDF: {path_pdf}")
        logger.info("=" * 60)

        return fond, scores, resultats_pdf, rapport

    def analyze(self, ticker: str) -> Optional[AnalysisResult]:
        self.fetch_fundamentals()

        df = self.connector.get_historique(ticker, days=730)
        if df.empty:
            logger.warning(f"{ticker}: pas de donnees historiques")
            return None

        df = self.history.save(ticker, df)

        if len(df) < 50:
            logger.warning(f"{ticker}: historique insuffisant ({len(df)} lignes)")
            return None

        live = self.connector.get_live(ticker) if self.connector.med_available else {}
        df = TechnicalIndicators.calculate_all(df)

        price = SafeMath.last(df["close"])
        fundamentals = self.fundamentals_cache.get(ticker.upper(), {})

        score_tech = TechnicalScorer().score(df)
        # PER dynamique
        if live:
            c = live.get("cours", 0) or 0
            b25 = fundamentals.get("bpa_2025", 0) or 0
            b26 = fundamentals.get("bpa_2026", 0) or 0
            if c > 0 and b25 > 0:
                fundamentals["per"] = round(c / b25, 1)
                fundamentals["forward_per"] = round(c / b26, 1) if b26 > 0 else fundamentals.get("forward_per")
        score_fond = FundamentalScorer(fundamentals).score()
        flags = RedFlagDetector(fundamentals).detect(df)
        setup = SetupDetector().detect(df)

        # Volume moyen pour ponderation dynamique
        vol_moy = df["volume"].tail(20).mean() * price if "volume" in df.columns else 0
        vr = SafeMath.last(df["Volume_Ratio"]) if "Volume_Ratio" in df.columns else 1.0

        # MASI filtre marché — bonus/malus selon régime
        masi_bonus = 0.0
        masi_regime = "Neutral"
        try:
            masi_data = self.connector.get_masi() if hasattr(self.connector, "get_masi") else {}
            masi_var = masi_data.get("variation", 0) or 0
            if masi_var > 0.5:
                masi_bonus = 0.3
                masi_regime = "Bullish"
            elif masi_var < -0.5:
                masi_bonus = -0.5
                masi_regime = "Risk-off"
        except:
            pass

        signal, action, score_global = DecisionEngine().decide(
            score_tech, score_fond, setup, flags, vol_moy, vr,
            fond=fundamentals, df=df
        )
        score_global = round(min(max(score_global + masi_bonus, 0), 10), 2)
        valuation = ValuationEngine(fundamentals).calculate(price, vr)

        result = AnalysisResult(
            ticker=ticker, price=price,
            date=MarketCalendar.now().replace(tzinfo=None),
            setup=setup, signal=signal, action=action,
            score_technical=score_tech, score_fundamental=score_fond,
            score_global=score_global, red_flags=flags,
            valuation=valuation, technical_data=score_tech.raw_details,
            fundamental_data=fundamentals, live_data=live,
            market_status=MarketCalendar.status()
        )
        result.institutional_note = NoteGenerator().generate(result)
        return result

    def run_full_analysis(self, tickers=None, backtest=False):
        """Analyse complete avec tous les modules"""
        tickers = tickers or TICKERS_DEFAUT

        print("=" * 80)
        print("BVC ANALYZER v5.2 — ANALYSE COMPLETE")
        print("RSI Memoire 10j + Fibonacci + Bonus Sectoriel + MASI Filtre")
        print("=" * 80)
        print(f"Heure Casablanca: {MarketCalendar.now().strftime('%H:%M:%S')}")
        print(f"Statut marche: {MarketCalendar.status()}")
        print(f"Message: {MarketCalendar.message()}")

        # 1. Test connectivite
        print("\nTest connecteurs...", end=" ")
        ok = self.connector.preload()
        print("✅ OK" if ok else "⚠️ Fallback uniquement")

        # 2. Pipeline PDF (optionnel — une fois par jour)
        fond, scores, pdfs, rapport = None, None, None, None
        try:
            fond, scores, pdfs, rapport = self.run_pdf_pipeline()
            self.fundamentals_cache = fond
        except Exception as e:
            logger.warning(f"Pipeline PDF saute: {e}")
            self.fetch_fundamentals()

        # 3. Analyse tickers
        print(f"\nAnalyse de {len(tickers)} titres...\n")
        results = {}
        backtests = {}

        for ticker in tickers:
            try:
                r = self.analyze(ticker)
                if r:
                    results[ticker] = r
                    print(f"  ✅ {ticker:6} | {r.price:>10.2f} DH | {r.action:<45} | Score {r.score_global}")
                    print(r.institutional_note)

                    # Backtest optionnel
                    if backtest and r.signal in [Signal.ACHAT_FORT, Signal.ACHETER, Signal.ACCUMULER]:
                        df = self.history.load(ticker)
                        if not df.empty and len(df) > 60:
                            # Signaux simples pour backtest: 1 si signal ACHAT, 0 sinon
                            signaux = pd.Series([False] * len(df))
                            signaux.iloc[-20:] = True  # Simuler 20j de signal
                            bt = BacktesterBVC().run(df, ticker, signaux)
                            backtests[ticker] = bt
                else:
                    print(f"  ❌ {ticker:6} | Donnees indisponibles")
            except Exception as e:
                print(f"  ❌ {ticker:6} | Erreur: {e}")

        # 4. Exports
        if results:
            rows = []
            for r in sorted(results.values(), key=lambda x: x.score_global, reverse=True):
                live = r.live_data or {}
                rows.append({
                    "Ticker": r.ticker, "Prix": r.price, "Signal": r.signal.value,
                    "Action": r.action, "Score Global": r.score_global,
                    "Technique": r.score_technical.normalized,
                    "Fondamental": r.score_fundamental.normalized,
                    "Setup": r.setup.value, "Red Flags": len(r.red_flags),
                    "Base": r.valuation.base, "Upside %": r.valuation.upside_base,
                    "Marche": r.market_status, "Heure": MarketCalendar.now().strftime("%H:%M"),
                    "Source": live.get("source", "Medias24"),
                    "Cours Live": live.get("cours", "N/A"),
                    "Variation": live.get("variation", "N/A"),
                })

            df_out = pd.DataFrame(rows)
            print("\n" + "=" * 80)
            print("CLASSEMENT GLOBAL")
            print("=" * 80)
            print(df_out.to_string(index=False))

            # Export Excel decision
            path_decision = self.export_engine.rapport_decision(list(results.values()))

            # Export JSON
            if fond:
                path_json = self.export_engine.export_json(fond)

            # Export synthese simple
            export = f"exports/bvc_v50_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            df_out.to_excel(export, index=False)
            logger.info(f"Export synthese: {export}")

            # Download Colab
            try:
                from google.colab import files
                files.download(export)
                files.download(path_decision)
            except: pass

        print("\n" + "=" * 80)
        print("ANALYSE TERMINEE")
        print("Historiques: data/historique/")
        print("Exports: exports/")
        if backtests:
            print(f"Backtests: {len(backtests)} strategies testees")
        print("=" * 80)

        # CLASSEMENTS MULTIPLES v5.2
        def top(lst, key, n=5, reverse=True):
            return sorted([r for r in lst if hasattr(r, 'signal') and r.signal not in ["EVITER"]], key=key, reverse=reverse)[:n]

        try:
            top_momentum  = top(results, lambda r: r.score_technical.normalized if r.score_technical else 0)
            top_fond      = top(results, lambda r: r.score_fundamental.normalized if r.score_fundamental else 0)
            top_global    = top(results, lambda r: r.score_global or 0)
            top_div       = top(results, lambda r: float((r.fundamental_data or {}).get("div_yield", 0) or 0))
            top_risque    = top(results, lambda r: -(len(r.red_flags or [])))

            print("=" * 60)
            print(" CLASSEMENTS MULTIPLES v5.2")
            print("=" * 60)

            for titre, lst in [
                ("🚀 TOP MOMENTUM",   top_momentum),
                ("📊 TOP FONDAMENTAL", top_fond),
                ("🏆 TOP GLOBAL",      top_global),
                ("💰 TOP DIVIDENDE",   top_div),
                ("🛡️  TOP RISQUE FAIBLE", top_risque),
            ]:
                print(titre)
                print("-" * 40)
                for i, r in enumerate(lst, 1):
                    sig = r.action[:30] if r.action else "-"
                    print(f"  {i}. {r.ticker:<6} {sig:<30} Score:{r.score_global}")
        except Exception as e:
            print(f"Classements erreur: {e}")

        return results, backtests, fond, scores

# ============================================================
# PUSH GITHUB (v3.5.3)
# ============================================================

def _push_github(file_path, token, repo="abdmoutalib207-lang/-bvc-analyzer", branch="main"):
    if not token: return
    fname = os.path.basename(file_path)
    hdrs = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}
    url = f"https://api.github.com/repos/{repo}/contents/{fname}"

    r = requests.get(url, headers=hdrs)
    sha = r.json().get("sha") if r.status_code == 200 else None

    with open(file_path, "rb") as f:
        content = base64.b64encode(f.read()).decode()

    payload = {
        "message": f"v5.0 enrichissement {datetime.today().strftime('%Y-%m-%d')}",
        "content": content, "branch": branch
    }
    if sha: payload["sha"] = sha

    r = requests.put(url, headers=hdrs, json=payload)
    if r.status_code in [200, 201]:
        logger.info("✅ Push GitHub reussi")
    else:
        logger.warning(f"❌ Push GitHub echoue: {r.status_code}")

# ============================================================
# MAIN
# ============================================================

def main():
    analyzer = BVCAnalyzer()
    results, backtests, fond, scores = analyzer.run_full_analysis(
        tickers=TICKERS_DEFAUT,
        backtest=False  # Mettre True pour activer les backtests
    )
    return results, backtests, fond, scores

if __name__ == "__main__":
    results, backtests, fond, scores = main()